import openpyxl
import re
wb = openpyxl.load_workbook("question-choosing.xlsx").active
dataDict = {}
resultDict = {}

for i in range(4, 29):      
    for j in range(2, 12, 4):
        key = wb.cell(row=i, column=j).value
        if (key != None):
            key = re.sub('\s+', '', key)
            key = key[-2:]
            dataDict[key] = wb.cell(row=i, column=j+1).value

groupResults = open("result.txt","w")
studentNames = open("../studentList.txt").readlines() 

for i in range(1,26):
    groupResults.write("\n\nQuestion "+str(i)+"\n")
    result = [studentNames[int(k)-1] for k,v in dataDict.items() if v==i]
    groupResults.writelines(result)

for i in resultDict.items():
    print(i)

